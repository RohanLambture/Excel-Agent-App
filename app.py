import streamlit as st
from dotenv import load_dotenv
import pandas as pd
import os
from langchain_groq import ChatGroq
from pandasai import SmartDataframe
from pandasai.llm import LLM
from io import BytesIO
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, LineChart, ScatterChart, Reference
from openpyxl.styles import Alignment

import base64

load_dotenv()

llm = ChatGroq(
    model_name="mixtral-8x7b-32768", 
    api_key=os.getenv("GROQ_API_KEY")
)

def process_query(df, prompt):
    smart_df = SmartDataframe(df, config={"llm": llm})
    
    response = smart_df.chat(prompt)
    
    if plt.get_fignums():
        fig = plt.gcf()
        buf = BytesIO()
        fig.savefig(buf, format="png")
        buf.seek(0)
        plt.close(fig)  
        return buf, "plot", smart_df.last_code_executed
    else:
        return response, "text", None

def extract_relevant_columns(code, df):
    columns = set(df.columns)
    used_columns = [col for col in columns if col in code]
    if not used_columns:
        return list(df.columns)
    if len(used_columns) == 1:
        return [used_columns[0]] + [col for col in df.columns if col not in used_columns][:1]
    return used_columns


def create_excel_with_plot_and_data(df, code):
    wb = Workbook()
    ws = wb.active


    ws['A1'] = "Code used to generate the plot:"
    ws['A2'] = code
    ws['A2'].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[2].height = 100  


    relevant_columns = extract_relevant_columns(code, df)
    relevant_df = df[relevant_columns]


    start_row = 4
    for r_idx, row in enumerate(dataframe_to_rows(relevant_df, index=False, header=True), start_row):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)


    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Sales by Year"
    chart.y_axis.title = 'Sales'
    chart.x_axis.title = 'Year'


    data = Reference(ws, min_col=2, min_row=start_row, max_row=start_row + len(relevant_df))
    cats = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + len(relevant_df))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    ws.add_chart(chart, "A" + str(start_row + len(relevant_df) + 2))

    excel_bytes = BytesIO()
    wb.save(excel_bytes)
    excel_bytes.seek(0)

    return excel_bytes

def verify_plot(plot_bytes, code, df):
    verification_bytes = BytesIO(plot_bytes.getvalue())
    plot_base64 = base64.b64encode(verification_bytes.getvalue()).decode()

    prompt = f"""
    Analyze the following plot and code:

    Plot (base64 encoded): {plot_base64}
    Code: {code}
    Data columns: {', '.join(df.columns)}

    Provide a very concise verification in the following format:
    1. Accuracy: [One sentence on plot accuracy]
    2. Issues: [List 1-2 main issues, if any, or "None" if no issues]
    3. Improvements: [List 1-2 key suggestions]

    Keep each point to one line, be extremely brief.
    """

    response = llm.invoke(prompt)
    return response

st.title("Excel Data Analysis and Visualization App")

uploaded_file = st.file_uploader("Choose an Excel file", type="xlsx")

if uploaded_file is not None:
    df = pd.read_excel(uploaded_file)
    
    st.write("Data Preview:")
    st.dataframe(df.head())
    
    prompt = st.text_input("Enter your query or plot prompt:", "")
    
    if st.button("Generate Response"):
        if prompt:
            result, result_type, code = process_query(df, prompt)
            
           
            if result_type == "plot":
                st.image(result, caption="Generated Plot", use_column_width=True)
                
                excel_bytes = create_excel_with_plot_and_data(df, code)
                st.download_button(
                    label="Download Plot and Data as Excel",
                    data=excel_bytes,
                    file_name="plot_and_data.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

                verification_result = verify_plot(result, code, df)
                st.subheader("Plot Verification:")
                st.text(verification_result)  
            else:
                st.write("Response:")
                st.write(result)
        else:
            st.write("Please enter a query or plot prompt.")
